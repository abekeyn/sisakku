# -*- coding: utf-8 -*-
"""鉄板焼きかいか（グラナダ）様 月次請求の【クラウド版】共通ロジック。

PC非依存。Supabase(Postgres)の出荷データから南志穂様宛の当月分を集計する。
GitHub Actions（月末スケジュール）とStreamlitアプリ（承認・送信）の双方から使う。

重要：注文テーブルには「旧シードの重複行(external_id空)」と
「ヤマト取込行(external_id='yamato:伝票番号')」が混在する月がある。
ヤマトの伝票番号が一意の正本なので、ヤマト取込行を優先し伝票で重複排除する。
"""
from __future__ import annotations

import calendar
import re
import subprocess
import unicodedata
from datetime import date
from pathlib import Path

import sqlalchemy as sa

from lib import db

CUSTOMER_ID = 7          # 南 志穂（鉄板焼き かいか）
UNIT_KG = 5.0
PRICE_PER_UNIT = 4000
TEMPLATE_XLSX = Path(__file__).resolve().parent.parent / "templates" / "granada_invoice_template.xlsx"
DOC_NO_KEY = "granada_last_doc_no"   # DB設定キー（最後に発番した書類番号）
DOC_NO_BASE = 260004                  # 2026年6月分(260004)まで発番済み
JP_DATE_FMT = 'yyyy"年"m"月"d"日"'


def _kg_from_name(name: str) -> float:
    t = unicodedata.normalize("NFKC", name or "").replace("kg", "㎏").replace("KG", "㎏")
    m = re.search(r"(\d+(?:\.\d+)?)\s*㎏", t)
    return float(m.group(1)) if m else 0.0


def month_shipments_db(target_ym: str, customer_id: int = CUSTOMER_ID) -> dict:
    """Supabaseから target_ym('YYYY-MM') の当該顧客宛出荷を集計。

    returns {"qty","total_kg","count","rows","source","warning"}
      rows: [{date, product, kg, denpyo}]
      source: 'yamato'(伝票あり) / 'seed'(伝票なし＝要注意)
      warning: 注意事項（あれば）
    """
    like = target_ym.replace("-", "/") + "%"   # '2026/07%'
    eng = db.get_engine()
    q = sa.text("""
        select o.id, o.ship_date, o.external_id, o.dispatch_ref,
               p.name as pname, p.weight_kg
        from orders o join products p on p.id = o.product_id
        where o.customer_id = :cid and o.ship_date like :like
        order by o.ship_date
    """)
    with eng.connect() as c:
        raw = [dict(r._mapping) for r in c.execute(q, {"cid": customer_id, "like": like})]

    # ヤマト取込行（external_id='yamato:...'）を伝票で一意化して優先採用
    yamato, seen = [], set()
    for r in raw:
        ext = (r.get("external_id") or "").strip()
        if ext.startswith("yamato:"):
            denpyo = ext.split(":", 1)[1]
            if denpyo in seen:
                continue
            seen.add(denpyo)
            yamato.append((r, denpyo))

    warning = ""
    if yamato:
        chosen = yamato
        source = "yamato"
    else:
        # ヤマト取込がまだ無い月：シード行を行IDで一意化（同期漏れの可能性）
        chosen = [(r, "") for r in raw]
        source = "seed"
        if raw:
            warning = "ヤマト取込データが未同期の可能性があります（伝票番号なし）。PCで最新の出荷取込をご確認ください。"

    rows, total_kg = [], 0.0
    for r, denpyo in chosen:
        kg = r["weight_kg"] or _kg_from_name(r["pname"])
        if kg <= 0:
            continue
        total_kg += kg
        rows.append({"date": r["ship_date"], "product": r["pname"],
                     "kg": kg, "denpyo": denpyo})
    return {"qty": total_kg / UNIT_KG, "total_kg": total_kg, "count": len(rows),
            "rows": rows, "source": source, "warning": warning}


def peek_next_doc_number() -> int:
    """次に発番する書類番号（DBの最後の番号+1）。確定はsend時にcommit_doc_number。"""
    last = db.get_setting(DOC_NO_KEY)
    try:
        last = int(last)
    except (TypeError, ValueError):
        last = DOC_NO_BASE
    return last + 1


def commit_doc_number(doc_no: int) -> None:
    db.set_setting(DOC_NO_KEY, int(doc_no))


def build_invoice_xlsx(target_ym: str, qty: float, doc_no: int,
                       out_path: str | Path,
                       template_path: str | Path = TEMPLATE_XLSX) -> dict:
    """テンプレートを openpyxl で複製し当月の値を埋めて保存（PC/Excel非依存）。

    可変セルのみ書き込み、書式は明示再設定（J1の和暦書式はopenpyxlが読み落とすため）。
    """
    from openpyxl import load_workbook

    y, m = int(target_ym[:4]), int(target_ym[5:7])
    last_day = calendar.monthrange(y, m)[1]
    issue = date(y, m, last_day)
    out_path = Path(out_path)

    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]
    ws.title = f"請求書 {target_ym.replace('-', '')}"
    ws["J1"] = issue
    ws["J1"].number_format = JP_DATE_FMT          # 「2026年7月31日」
    ws["J2"] = doc_no
    ws["B9"] = f"下記の通り、{m}月分をご請求申し上げます。"
    ws["F18"] = qty
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return {"path": str(out_path), "issue_date": issue.isoformat(),
            "amount": int(round(qty * PRICE_PER_UNIT)), "doc_number": doc_no,
            "sheet": ws.title}


def xlsx_to_pdf(xlsx_path: str | Path, out_dir: str | Path,
                soffice: str = "soffice") -> Path:
    """LibreOffice headless で xlsx→PDF（GitHub Actions/Linux用）。"""
    xlsx_path = Path(xlsx_path)
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    subprocess.run(
        [soffice, "--headless", "--calc", "--convert-to", "pdf",
         "--outdir", str(out_dir), str(xlsx_path)],
        check=True, timeout=120,
        stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    pdf = out_dir / (xlsx_path.stem + ".pdf")
    if not pdf.exists():
        raise RuntimeError(f"PDF生成に失敗: {pdf}")
    return pdf


def pdf_filename(issue_date_iso: str) -> str:
    return f"{issue_date_iso.replace('-', '')}_鉄板焼きかいか様_請求書.pdf"


# ===== 承認フロー（prepare = Actions側 / send = アプリ側） ==========================
import base64 as _b64  # noqa: E402
import json as _json   # noqa: E402
import urllib.request  # noqa: E402
from datetime import datetime  # noqa: E402

from lib import config  # noqa: E402

CUSTOMER_EMAIL = "keiri@granada-jp.net"
PENDING_KEY = "granada_pending_invoice"   # DB設定キー（承認待ち1件）


def _ntfy(title_ascii: str, body_jp: str, priority: str = "default",
          tags: str = "rice", click: str = "", actions: str = "") -> None:
    topic = config.get_secret("NTFY_TOPIC", "abe-rice-farm-claude")
    headers = {"Title": title_ascii, "Priority": priority, "Tags": tags,
               "Content-Type": "text/plain; charset=utf-8"}
    if click:
        headers["Click"] = click
    if actions:
        headers["Actions"] = actions
    try:
        urllib.request.urlopen(urllib.request.Request(
            f"https://ntfy.sh/{topic}", data=body_jp.encode("utf-8"),
            headers=headers, method="POST"), timeout=8)
    except Exception as e:  # noqa: BLE001
        print("ntfy通知失敗（継続）:", e)


def get_pending() -> dict | None:
    return db.get_setting(PENDING_KEY)


def prepare(target_ym: str | None = None, soffice: str = "soffice",
            workdir: str | Path = ".") -> dict:
    """当月分を集計→xlsx→PDF生成し、承認待ちとしてDBに保存、スマホへ承認通知。
    送信はしない（アプリの『送信する』で確定）。"""
    if target_ym is None:
        target_ym = datetime.now().strftime("%Y-%m")
    y, m = int(target_ym[:4]), int(target_ym[5:7])
    workdir = Path(workdir)

    s = month_shipments_db(target_ym)
    qty, amount = s["qty"], int(round(s["qty"] * PRICE_PER_UNIT))

    if s["count"] == 0 or qty <= 0:
        _ntfy("Granada invoice: NEEDS CHECK",
              f"⚠️ {m}月分の南志穂様宛発送が0件でした（{s['warning']}）。\n"
              f"請求書は作成していません。手動でご確認ください。",
              priority="high", tags="warning")
        return {"target_ym": target_ym, "count": 0, "status": "no_shipments"}

    doc_no = peek_next_doc_number()
    xlsx = workdir / f"granada_{target_ym}.xlsx"
    info = build_invoice_xlsx(target_ym, qty, doc_no, xlsx)
    pdf = xlsx_to_pdf(xlsx, workdir, soffice=soffice)
    pdf_b64 = _b64.b64encode(pdf.read_bytes()).decode("ascii")

    pending = {
        "target_ym": target_ym, "month": m, "issue_date": info["issue_date"],
        "doc_number": doc_no, "qty": qty, "total_kg": s["total_kg"],
        "amount": amount, "rows": s["rows"], "source": s["source"],
        "warning": s["warning"], "pdf_b64": pdf_b64,
        "pdf_name": pdf_filename(info["issue_date"]),
        "status": "pending", "prepared_at": datetime.now().isoformat(timespec="seconds"),
    }
    db.set_setting(PENDING_KEY, pending)

    app_url = config.get_secret("APP_URL", "")
    # ntfyのヘッダ(Click/Actions)はlatin-1のみ→URL・ラベルはASCIIで。本文は日本語OK。
    click = f"{app_url.rstrip('/')}/invoice_approve?invoice={target_ym}" if app_url else ""
    actions = (f"view, Review and Send, {click}, clear=true" if click else "")
    warn = f"\n⚠️ {s['warning']}" if s["warning"] else ""
    _ntfy("Granada invoice: APPROVE?",
          f"📨 {m}月分の請求書ができました。送信してよろしいですか？\n"
          f"金額 ¥{amount:,}（{qty:g}個 / {s['total_kg']:g}kg / {s['count']}件）\n"
          f"宛先 {CUSTOMER_EMAIL}{warn}\n"
          f"→ アプリで内容を確認し『送信する』を押すと送信します。",
          priority="high", tags="rice,email", click=click, actions=actions)
    return {**{k: pending[k] for k in ("target_ym", "amount", "qty", "doc_number")},
            "status": "prepared"}


def _smtp_send(subject: str, body: str, pdf_bytes: bytes, filename: str,
               to_addr: str = CUSTOMER_EMAIL) -> tuple[bool, str]:
    import smtplib
    import ssl
    from email.message import EmailMessage
    host = config.get_secret("SMTP_HOST", "smtp.gmail.com")
    port = int(config.get_secret("SMTP_PORT", "587") or 587)
    user = config.get_secret("SMTP_USER", "")
    pw = config.get_secret("SMTP_PASS", "")
    sender = config.get_secret("MAIL_FROM", "") or user
    if not (host and user and pw):
        return False, "SMTP未設定（SMTP_USER / SMTP_PASS）"
    msg = EmailMessage()
    msg["From"] = sender
    msg["To"] = to_addr
    msg["Subject"] = subject
    msg.set_content(body)
    msg.add_attachment(pdf_bytes, maintype="application", subtype="pdf",
                       filename=filename)
    try:
        ctx = ssl.create_default_context()
        if port == 465:
            with smtplib.SMTP_SSL(host, port, context=ctx, timeout=40) as sv:
                sv.login(user, pw)
                sv.send_message(msg)
        else:
            with smtplib.SMTP(host, port, timeout=40) as sv:
                sv.starttls(context=ctx)
                sv.login(user, pw)
                sv.send_message(msg)
        return True, f"送信しました（{to_addr}）"
    except Exception as e:  # noqa: BLE001
        return False, f"メール送信に失敗: {e}"


def _mail_text(year: int, month: int) -> tuple[str, str]:
    subject = f"阿部農園　{month}月分請求につきまして"
    body = ("株式会社グラナダ\n経理部\nご担当者様\n\n"
            "いつもお世話になっております。\n阿部農園の阿部と申します。\n\n"
            f"{year}年{month}月分の請求書を送付いたします。\n"
            "内容についてご確認の上、\nご不明点等ございましたらご返信をお願いいたします。\n\n"
            "以上、よろしくお願いいたします。\n")
    return subject, body


def send_pending() -> dict:
    """承認待ちの請求書を実際に顧客へ送信し、確定（書類番号commit・履歴記録）。"""
    p = get_pending()
    if not p:
        return {"ok": False, "msg": "承認待ちの請求書がありません"}
    if p.get("status") == "sent":
        return {"ok": False, "msg": "既に送信済みです", "pending": p}
    y, m = int(p["target_ym"][:4]), p["month"]
    subject, body = _mail_text(y, m)
    pdf = _b64.b64decode(p["pdf_b64"])
    ok, msg = _smtp_send(subject, body, pdf, p["pdf_name"])
    if not ok:
        _ntfy("Granada invoice: SEND FAILED",
              f"⚠️ {m}月分の送信に失敗しました。\n{msg}", priority="high", tags="warning")
        return {"ok": False, "msg": msg}
    commit_doc_number(p["doc_number"])
    p["status"] = "sent"
    p["sent_at"] = datetime.now().isoformat(timespec="seconds")
    p["synced_to_xlsx"] = False           # PCのagentが台帳へ追記したらTrue
    db.set_setting(PENDING_KEY, p)
    hist = db.get_setting("granada_sent_history") or []
    hist.append({k: p[k] for k in ("target_ym", "doc_number", "amount", "qty", "sent_at")})
    db.set_setting("granada_sent_history", hist)
    _ntfy("Granada invoice: SENT",
          f"✅ {m}月分の請求書を送信しました。\n金額 ¥{p['amount']:,}（{p['qty']:g}個）\n"
          f"宛先 {CUSTOMER_EMAIL}\n書類番号 {p['doc_number']}", tags="white_check_mark")
    return {"ok": True, "msg": msg, "pending": p}


def main() -> None:
    import argparse
    ap = argparse.ArgumentParser(description="グラナダ請求 クラウド版")
    ap.add_argument("--prepare", action="store_true", help="集計→PDF→承認待ち登録→通知")
    ap.add_argument("--send", action="store_true", help="承認待ちを送信（通常はアプリから）")
    ap.add_argument("--month", help="対象月 YYYY-MM（既定: 当月）")
    ap.add_argument("--soffice", default="soffice", help="LibreOffice実行パス")
    ap.add_argument("--workdir", default=".", help="一時ファイル出力先")
    a = ap.parse_args()
    try:
        if a.send:
            print(send_pending())
        else:
            print(prepare(a.month, soffice=a.soffice, workdir=a.workdir))
    except Exception as e:  # noqa: BLE001
        import traceback
        traceback.print_exc()
        _ntfy("Granada invoice: ERROR",
              f"❌ クラウド月次処理でエラー: {type(e).__name__}: {e}\n手動でご対応ください。",
              priority="urgent", tags="rotating_light")
        raise


if __name__ == "__main__":
    main()


def sync_local_xlsx() -> dict:
    """【PC側】クラウドで送信済みだがローカル台帳に未反映の請求書を追記する。
    agent.py（常駐）から呼ぶ。Excel COMを使うのでPC専用。"""
    p = get_pending()
    if not p or p.get("status") != "sent" or p.get("synced_to_xlsx"):
        return {"synced": False, "reason": "対象なし"}
    from lib import granada_invoice as gi  # PC専用（win32com）
    info = gi.generate(p["target_ym"], p["qty"], doc_no=p["doc_number"])
    p["synced_to_xlsx"] = True
    p["synced_at"] = datetime.now().isoformat(timespec="seconds")
    db.set_setting(PENDING_KEY, p)
    print("グラナダ台帳へ追記:", info["sheet_name"], info["pdf_path"], flush=True)
    return {"synced": True, "sheet": info["sheet_name"], "pdf": info["pdf_path"]}
