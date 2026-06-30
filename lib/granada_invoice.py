# -*- coding: utf-8 -*-
"""鉄板焼きかいか（株式会社グラナダ）様 月次請求書の自動作成。

南志穂様宛の当月発送（ヤマト発行済データ）を集計し、5kgあたり送料込み4000円で
請求書シートを「グラナダ様請求書.xlsx」に追加、当該シートをPDF出力する。

請求ロジック（2026年4月以降の運用形式に準拠）:
    当月の総kg ÷ 5 = 数量(5kg換算の個数)、金額(税込) = 数量 × 4000、8%軽減税率。
    （検証: 4月=55kg→11個→44,000円 / 5月=20kg→4個→16,000円 と一致）

メール送信・スマホ通知はこのモジュールでは行わない（呼び出し側＝Claude/スケジュール
ジョブが mailer 経由で行う）。ここは「数字の確定」と「帳票生成」までを担う。
"""
from __future__ import annotations

import calendar
import csv
import io
import re
import unicodedata
from datetime import date
from pathlib import Path

INVOICE_XLSX = Path(
    r"C:/Users/wolhp/OneDrive/デスクトップ/発行書類/鉄板焼きかいか様/グラナダ様請求書.xlsx"
)
OUT_DIR = INVOICE_XLSX.parent
UNIT_KG = 5.0            # 1個あたり
PRICE_PER_UNIT = 4000   # 税込・送料込み（5kgあたり）
TAX_RATE = 0.08         # 軽減税率
RECIPIENT_KEYS = ("南", "志穂")  # お届け先名の判定
RECIPIENT_ZIP = "1040061"        # 念のための補助判定（南志穂様）


def _z2h(s: str) -> str:
    """全角英数字を半角へ。"品名"の "10㎏"/"1５㎏" 等を数値化できるように。"""
    return unicodedata.normalize("NFKC", s or "")


def kg_from_hinmei(hinmei: str) -> float:
    """品名（例: '精米５㎏', '精米10㎏', '精米1５㎏'）からkg数を取り出す。
    取り出せない（やさい等）場合は 0.0。"""
    t = _z2h(hinmei).replace("kg", "㎏").replace("KG", "㎏")
    m = re.search(r"(\d+(?:\.\d+)?)\s*㎏", t)
    return float(m.group(1)) if m else 0.0


def _is_recipient(name: str, zipcode: str = "") -> bool:
    n = name or ""
    if all(k in n for k in RECIPIENT_KEYS):
        return True
    z = re.sub(r"\D", "", zipcode or "")
    return bool(z) and z == RECIPIENT_ZIP


def summarize_month(issued_csv: bytes, target_ym: str) -> dict:
    """ヤマト発行済CSV(bytes, cp932)から target_ym('YYYY-MM') の南志穂様宛を集計。

    returns {"qty", "total_kg", "count", "rows"}
      qty: 5kg換算の個数（= total_kg / 5）
      rows: [{date, hinmei, kg, tracking}] 明細（確認用）
    """
    text = issued_csv.decode("cp932", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return {"qty": 0.0, "total_kg": 0.0, "count": 0, "rows": []}
    hdr = rows[0]

    def col(name: str, default: int) -> int:
        try:
            return hdr.index(name)
        except ValueError:
            return default

    c_date = col("出荷予定日", 4)
    c_name = col("お届け先名", 15)
    c_zip = col("お届け先郵便番号", 10)
    c_hin = col("品名１", 27)
    c_trk = col("伝票番号", 3)

    ym_prefix = target_ym.replace("-", "/")  # '2026/07'
    detail = []
    total_kg = 0.0
    for r in rows[1:]:
        if len(r) <= max(c_date, c_name, c_hin):
            continue
        d = (r[c_date] or "").strip().replace("-", "/")
        if not d.startswith(ym_prefix):
            continue
        if not _is_recipient(r[c_name], r[c_zip] if len(r) > c_zip else ""):
            continue
        kg = kg_from_hinmei(r[c_hin])
        if kg <= 0:
            continue
        total_kg += kg
        detail.append({"date": d, "hinmei": r[c_hin].strip(),
                       "kg": kg, "tracking": (r[c_trk] if len(r) > c_trk else "").strip()})
    return {"qty": total_kg / UNIT_KG, "total_kg": total_kg,
            "count": len(detail), "rows": detail}


def pdf_name(issue_date_iso: str) -> str:
    return f"{issue_date_iso.replace('-', '')}_鉄板焼きかいか様_請求書.pdf"


def _peek_sheets(xlsx_path: Path) -> tuple[list[str], int]:
    """既存シート名一覧と、次に使う書類番号を openpyxl で取得（COMより速く安全）。"""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True)
    names = list(wb.sheetnames)
    nums = []
    for sn in names:
        v = wb[sn]["J2"].value
        if isinstance(v, (int, float)):
            nums.append(int(v))
    wb.close()
    return names, (max(nums) + 1 if nums else 260001)


def generate(target_ym: str, qty: float, xlsx_path: Path = INVOICE_XLSX,
             out_dir: Path = OUT_DIR, doc_no: int | None = None) -> dict:
    """target_ym('YYYY-MM') の請求書シートを追加し、当該シートをPDF出力する。

    既存の最新シートを Excel でそのまま複製するため、書式・結合・印刷設定を完全継承。
    冪等: 当月シートが既にあれば複製せず、値を当月の確定値に更新してPDFを出し直す。
    doc_no を渡すと書類番号をそれに固定（クラウド発番との整合用）。

    返り値 {"sheet_name","doc_number","issue_date","amount","pdf_path","created"}
    """
    import win32com.client as win32

    y, m = int(target_ym[:4]), int(target_ym[5:7])
    last_day = calendar.monthrange(y, m)[1]
    issue_dt = date(y, m, last_day)
    issue_serial = (issue_dt - date(1899, 12, 30)).days  # Excelシリアル値
    new_name = f"請求書 {target_ym.replace('-', '')}"  # 例: 請求書 202607
    amount = int(round(qty * PRICE_PER_UNIT))

    names, next_no = _peek_sheets(xlsx_path)
    if doc_no is not None:
        next_no = int(doc_no)
    existed = new_name in names
    template = None
    if not existed:
        inv = [s for s in names if "請求" in s and re.search(r"20\d{4}", s)]
        inv.sort(key=lambda s: re.search(r"20\d{4}", s).group(0))
        if not inv:
            raise RuntimeError("テンプレートとなる既存請求書シートが見つかりません")
        template = inv[-1]

    out_pdf = out_dir / pdf_name(issue_dt.isoformat())
    out_dir.mkdir(parents=True, exist_ok=True)

    # 早期バインディング(EnsureDispatch)。Worksheet.Copy の名前付き引数を正しく束ねる。
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(str(xlsx_path))
        if existed:
            ws = wb.Worksheets(new_name)
            doc_no = int(ws.Range("J2").Value)  # 採番済みは維持
        else:
            src = wb.Worksheets(template)
            src.Copy(After=wb.Worksheets(wb.Worksheets.Count))  # 末尾に完全複製
            ws = wb.Worksheets(wb.Worksheets.Count)
            ws.Name = new_name
            doc_no = next_no
            ws.Range("J2").Value = doc_no
        # 当月の確定値を書き込む（書式は複製元を継承）
        # 日付はシリアル値で設定（datetime渡しだとCOMのTZ変換で1日ずれるため）
        ws.Range("J1").Value2 = issue_serial                 # 発行日（末日）
        ws.Range("B9").Value = f"下記の通り、{m}月分をご請求申し上げます。"
        ws.Range("F18").Value = qty                          # 数量（5kg換算個数）
        wb.Save()
        ws.ExportAsFixedFormat(0, str(out_pdf))              # 0=PDF, このシートのみ
        wb.Close(SaveChanges=False)
    finally:
        excel.Quit()

    return {"sheet_name": new_name, "doc_number": doc_no,
            "issue_date": issue_dt.isoformat(), "amount": amount,
            "pdf_path": str(out_pdf), "created": not existed}
