# -*- coding: utf-8 -*-
"""請求（月次）汎用エンジン — 複数の請求先(取引先)に対応。

請求先マスタ（settingsにJSONで保持）ごとに、
  - 集計対象の顧客(customer_id)の当月出荷を伝票番号で重複排除して集計
  - 5kgあたりの単価で金額算出
  - テンプレートから請求書xlsx→PDFを生成
  - 承認待ちとして保存し、スマホへ承認通知
  - アプリで承認→定型文でメール送信→書類番号を確定
を行う。PC非依存（GitHub Actions＋Streamlitアプリ）。

スキーマ変更不要：マスタも承認待ちもすべて settings テーブルにJSONで保存する。
"""
from __future__ import annotations

import base64
import calendar
import re
from datetime import date, datetime
from pathlib import Path

import sqlalchemy as sa

from lib import config, db
from lib.granada_cloud import (JP_DATE_FMT, TEMPLATE_XLSX, _kg_from_name,
                               _ntfy, _smtp_send, pdf_filename, xlsx_to_pdf)

CLIENTS_KEY = "billing_clients"
PENDING_KEY = "billing_pending"          # {f"{client_id}:{ym}": pending}
HISTORY_KEY = "billing_sent_history"
UNIT_KG = 5.0

# 既定の請求先（初回はこれをマスタへ取り込む）。鉄板焼きかいか（グラナダ）様。
DEFAULT_CLIENTS = [{
    "id": "granada",
    "name": "鉄板焼きかいか（グラナダ）様",
    "invoice_to": "株式会社グラナダ 鉄板焼きかいか",
    "email": "keiri@granada-jp.net",
    "customer_id": 7,
    "price_per_5kg": 4000,
    "item_desc": "令和7年度　福島県産 コシヒカリ (精米) 5㎏",
    "subject_tmpl": "阿部農園　{month}月分請求につきまして",
    "body_tmpl": (
        "株式会社グラナダ\n経理部\nご担当者様\n\n"
        "いつもお世話になっております。\n阿部農園の阿部と申します。\n\n"
        "{year}年{month}月分の請求書を送付いたします。\n"
        "内容についてご確認の上、\nご不明点等ございましたらご返信をお願いいたします。\n\n"
        "以上、よろしくお願いいたします。\n"),
    "last_doc_no": 260004,
    "local_xlsx": r"C:\Users\wolhp\OneDrive\デスクトップ\発行書類\鉄板焼きかいか様\グラナダ様請求書.xlsx",
    "active": True,
}]


# ===== マスタCRUD ===================================================================
def get_clients() -> list[dict]:
    cs = db.get_setting(CLIENTS_KEY)
    if not cs:
        db.set_setting(CLIENTS_KEY, DEFAULT_CLIENTS)
        return [dict(c) for c in DEFAULT_CLIENTS]
    return cs


def get_client(client_id: str) -> dict | None:
    return next((c for c in get_clients() if c["id"] == client_id), None)


def save_clients(clients: list[dict]) -> None:
    db.set_setting(CLIENTS_KEY, clients)


def upsert_client(client: dict) -> None:
    cs = get_clients()
    for i, c in enumerate(cs):
        if c["id"] == client["id"]:
            cs[i] = client
            break
    else:
        cs.append(client)
    save_clients(cs)


def delete_client(client_id: str) -> None:
    save_clients([c for c in get_clients() if c["id"] != client_id])


# ===== 集計 =======================================================================
def month_shipments(client: dict, target_ym: str) -> dict:
    """請求先clientの集計対象顧客の、target_ym('YYYY-MM')出荷を伝票で重複排除集計。"""
    like = target_ym.replace("-", "/") + "%"
    q = sa.text("""
        select o.id, o.ship_date, o.external_id, p.name as pname, p.weight_kg
        from orders o join products p on p.id = o.product_id
        where o.customer_id = :cid and o.ship_date like :like
        order by o.ship_date
    """)
    with db.get_engine().connect() as c:
        raw = [dict(r._mapping) for r in
               c.execute(q, {"cid": client["customer_id"], "like": like})]

    yamato, seen = [], set()
    for r in raw:
        ext = (r.get("external_id") or "").strip()
        if ext.startswith("yamato:"):
            d = ext.split(":", 1)[1]
            if d in seen:
                continue
            seen.add(d)
            yamato.append((r, d))
    warning = ""
    if yamato:
        chosen, source = yamato, "yamato"
    else:
        chosen, source = [(r, "") for r in raw], "seed"
        if raw:
            warning = "ヤマト取込データが未同期の可能性があります（伝票番号なし）。PCで最新の出荷取込をご確認ください。"

    rows, total_kg = [], 0.0
    for r, d in chosen:
        kg = r["weight_kg"] or _kg_from_name(r["pname"])
        if kg <= 0:
            continue
        total_kg += kg
        rows.append({"date": r["ship_date"], "product": r["pname"], "kg": kg, "denpyo": d})
    price = client.get("price_per_5kg", 4000)
    qty = total_kg / UNIT_KG
    return {"qty": qty, "total_kg": total_kg, "count": len(rows), "rows": rows,
            "amount": int(round(qty * price)), "source": source, "warning": warning}


# ===== 帳票生成 ===================================================================
def build_invoice_xlsx(client: dict, target_ym: str, qty: float, doc_no: int,
                       out_path: str | Path) -> dict:
    """テンプレートを複製し、請求先固有の宛名・品名・当月値を埋めて保存。"""
    from openpyxl import load_workbook
    y, m = int(target_ym[:4]), int(target_ym[5:7])
    issue = date(y, m, calendar.monthrange(y, m)[1])
    out_path = Path(out_path)
    wb = load_workbook(TEMPLATE_XLSX)
    ws = wb[wb.sheetnames[0]]
    ws.title = f"請求書 {target_ym.replace('-', '')}"
    ws["B5"] = client["invoice_to"]                       # 宛名
    ws["B18"] = client.get("item_desc", ws["B18"].value)  # 品名
    ws["J1"] = issue
    ws["J1"].number_format = JP_DATE_FMT
    ws["J2"] = doc_no
    ws["B9"] = f"下記の通り、{m}月分をご請求申し上げます。"
    ws["F18"] = qty
    # LibreOffice変換で1ページに収める（Excelと違い自動では収まらないため明示）
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.print_area = "A1:J32"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return {"issue_date": issue.isoformat(), "sheet": ws.title}


# ===== 承認フロー =================================================================
def get_pendings() -> dict:
    return db.get_setting(PENDING_KEY) or {}


def _save_pendings(p: dict) -> None:
    db.set_setting(PENDING_KEY, p)


def prepare_client(client: dict, target_ym: str, soffice: str = "soffice",
                   workdir: str | Path = ".") -> dict:
    """1請求先の当月分を作成し承認待ちに登録。返り値はpendingの要約。"""
    y, m = int(target_ym[:4]), int(target_ym[5:7])
    s = month_shipments(client, target_ym)
    if s["count"] == 0 or s["qty"] <= 0:
        return {"client_id": client["id"], "status": "no_shipments",
                "name": client["name"], "warning": s["warning"]}

    doc_no = int(client.get("last_doc_no", 0)) + 1
    workdir = Path(workdir)
    xlsx = workdir / f"invoice_{client['id']}_{target_ym}.xlsx"
    info = build_invoice_xlsx(client, target_ym, s["qty"], doc_no, xlsx)
    pdf = xlsx_to_pdf(xlsx, workdir, soffice=soffice)
    pending = {
        "client_id": client["id"], "client_name": client["name"],
        "email": client["email"], "target_ym": target_ym, "month": m,
        "issue_date": info["issue_date"], "doc_number": doc_no,
        "qty": s["qty"], "total_kg": s["total_kg"], "amount": s["amount"],
        "rows": s["rows"], "source": s["source"], "warning": s["warning"],
        "pdf_b64": base64.b64encode(pdf.read_bytes()).decode("ascii"),
        "pdf_name": pdf_filename(info["issue_date"]),
        "status": "pending", "prepared_at": datetime.now().isoformat(timespec="seconds"),
    }
    allp = get_pendings()
    allp[f"{client['id']}:{target_ym}"] = pending
    _save_pendings(allp)
    return {"client_id": client["id"], "name": client["name"],
            "amount": s["amount"], "qty": s["qty"], "status": "prepared"}


def prepare_all(target_ym: str | None = None, soffice: str = "soffice",
                workdir: str | Path = ".") -> dict:
    """全アクティブ請求先の当月分を作成し、まとめてスマホへ承認通知。"""
    if target_ym is None:
        target_ym = datetime.now().strftime("%Y-%m")
    m = int(target_ym[5:7])
    prepared, skipped = [], []
    for c in get_clients():
        if not c.get("active", True):
            continue
        r = prepare_client(c, target_ym, soffice=soffice, workdir=workdir)
        (prepared if r["status"] == "prepared" else skipped).append(r)

    app_url = config.get_secret("APP_URL", "")
    click = f"{app_url.rstrip('/')}/?tab=billing" if app_url else ""
    actions = f"view, Open app, {click}, clear=true" if click else ""
    if prepared:
        lines = "\n".join(f"・{p['name']} ¥{p['amount']:,}" for p in prepared)
        warn = ("\n⚠️ 0件/要確認: " + "、".join(s["name"] for s in skipped)) if skipped else ""
        _ntfy("Invoice: APPROVE?",
              f"📨 {m}月分の請求書ができました（{len(prepared)}件）。送信してよろしいですか？\n"
              f"{lines}{warn}\n→ アプリで内容を確認し『送信する』を押してください。",
              priority="high", tags="rice,email", click=click, actions=actions)
    elif skipped:
        _ntfy("Invoice: NEEDS CHECK",
              f"⚠️ {m}月分は送信対象がありませんでした（{len(skipped)}件 要確認）。\n"
              + "、".join(s["name"] for s in skipped), priority="high", tags="warning")
    return {"target_ym": target_ym, "prepared": prepared, "skipped": skipped}


def send_pending(pending_key: str) -> dict:
    """承認待ち1件を顧客へ送信し、書類番号を確定。pending_key='client:ym'。"""
    allp = get_pendings()
    p = allp.get(pending_key)
    if not p:
        return {"ok": False, "msg": "対象の請求書が見つかりません"}
    if p.get("status") == "sent":
        return {"ok": False, "msg": "既に送信済みです"}
    client = get_client(p["client_id"])
    if not client:
        return {"ok": False, "msg": "請求先マスタが見つかりません"}
    y, m = int(p["target_ym"][:4]), p["month"]
    subject = client["subject_tmpl"].format(year=y, month=m)
    body = client["body_tmpl"].format(year=y, month=m)
    ok, msg = _smtp_send(subject, body, base64.b64decode(p["pdf_b64"]),
                         p["pdf_name"], to_addr=p["email"])
    if not ok:
        _ntfy("Invoice: SEND FAILED",
              f"⚠️ {p['client_name']} {m}月分の送信に失敗：{msg}",
              priority="high", tags="warning")
        return {"ok": False, "msg": msg}
    # 書類番号を確定（マスタのlast_doc_noを更新）
    client["last_doc_no"] = p["doc_number"]
    upsert_client(client)
    p["status"] = "sent"
    p["sent_at"] = datetime.now().isoformat(timespec="seconds")
    p["synced_to_xlsx"] = False
    allp[pending_key] = p
    _save_pendings(allp)
    hist = db.get_setting(HISTORY_KEY) or []
    hist.append({k: p[k] for k in ("client_id", "client_name", "target_ym",
                                   "doc_number", "amount", "sent_at")})
    db.set_setting(HISTORY_KEY, hist)
    _ntfy("Invoice: SENT",
          f"✅ {p['client_name']} {m}月分を送信しました。\n"
          f"金額 ¥{p['amount']:,} / 宛先 {p['email']} / 書類番号 {p['doc_number']}",
          tags="white_check_mark")
    return {"ok": True, "msg": msg, "pending": p}


def discard_pending(pending_key: str) -> None:
    allp = get_pendings()
    if pending_key in allp:
        del allp[pending_key]
        _save_pendings(allp)


# ===== PC側：ローカルExcel台帳へ同期追記 =========================================
def sync_local_xlsx() -> list[dict]:
    """送信済みだがローカル台帳未反映の請求書を、各請求先のlocal_xlsxへ追記。"""
    out = []
    allp = get_pendings()
    changed = False
    for key, p in allp.items():
        if p.get("status") != "sent" or p.get("synced_to_xlsx"):
            continue
        client = get_client(p["client_id"]) or {}
        xlsx = client.get("local_xlsx")
        if not xlsx or not Path(xlsx).exists():
            continue  # ローカル台帳が無い請求先はPDFのみ（クラウド保管）
        from lib import granada_invoice as gi  # win32com（PC専用）
        info = gi.generate(p["target_ym"], p["qty"], xlsx_path=Path(xlsx),
                           out_dir=Path(xlsx).parent, doc_no=p["doc_number"])
        p["synced_to_xlsx"] = True
        p["synced_at"] = datetime.now().isoformat(timespec="seconds")
        changed = True
        out.append({"client": p["client_name"], "sheet": info["sheet_name"]})
    if changed:
        _save_pendings(allp)
    return out


def main() -> None:
    import argparse
    ap = argparse.ArgumentParser(description="月次請求 クラウド版（複数請求先）")
    ap.add_argument("--prepare", action="store_true")
    ap.add_argument("--month")
    ap.add_argument("--soffice", default="soffice")
    ap.add_argument("--workdir", default=".")
    a = ap.parse_args()
    try:
        print(prepare_all(a.month, soffice=a.soffice, workdir=a.workdir))
    except Exception as e:  # noqa: BLE001
        import traceback
        traceback.print_exc()
        _ntfy("Invoice: ERROR",
              f"❌ クラウド月次処理でエラー: {type(e).__name__}: {e}",
              priority="urgent", tags="rotating_light")
        raise


if __name__ == "__main__":
    main()
